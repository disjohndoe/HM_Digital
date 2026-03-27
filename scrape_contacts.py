"""
Scrape clinic contacts from mojausluga.hr and merge with HZJZ Excel.
Outputs: docs/medical_leads.csv

Usage: python -X utf8 scrape_contacts.py
"""

import csv
import sys
import time
import requests
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

HZJZ_EXCEL = Path.home() / "Downloads" / "Popis_pruzatelja_NRPZZ_03.11.2025..xlsx"
OUTPUT_CSV = Path(__file__).parent / "docs" / "medical_leads.csv"
BASE_URL = "https://mojausluga.hr/zubari"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Accept": "application/json"}
DENTAL_KEYWORDS = ["dent", "stomat", "zub", "oral", "ortodon", "endodont", "parodont", "protet"]


def is_dental(name: str) -> bool:
    return any(kw in name.lower() for kw in DENTAL_KEYWORDS)


def parse_city(slug: str) -> list[dict]:
    """Fetch and parse all clinics for a city from SvelteKit __data.json."""
    try:
        resp = requests.get(f"{BASE_URL}/{slug}/__data.json", headers=HEADERS, timeout=15)
        resp.raise_for_status()
        nodes = resp.json().get("nodes", [])
    except Exception as e:
        return []

    for node in nodes:
        if not isinstance(node, dict) or "data" not in node:
            continue
        data = node["data"]
        if not isinstance(data, list) or len(data) < 20:
            continue

        # Find the providers list (index 9 in the standard layout)
        # Verify by checking that data[9] is a list of integers
        provider_list = None
        for i in range(7, 12):
            if i < len(data) and isinstance(data[i], list) and data[i] and isinstance(data[i][0], int):
                provider_list = data[i]
                break

        if not provider_list:
            continue

        clinics = []
        for si in provider_list:
            if si >= len(data):
                continue
            entry = data[si]
            if not isinstance(entry, dict):
                continue

            c = {"city_slug": slug}
            for field, idx in entry.items():
                if isinstance(idx, int) and idx < len(data):
                    v = data[idx]
                    if isinstance(v, (str, int, float, bool)):
                        c[field] = v

            # Also try to extract doctor info
            doctors_field = entry.get("doctors")
            if isinstance(doctors_field, list) and doctors_field:
                first_doc_idx = doctors_field[0]
                if isinstance(first_doc_idx, int) and first_doc_idx < len(data):
                    doc_entry = data[first_doc_idx]
                    if isinstance(doc_entry, dict):
                        for field, idx in doc_entry.items():
                            if isinstance(idx, int) and idx < len(data):
                                v = data[idx]
                                if isinstance(v, str):
                                    c[f"doctor_{field}"] = v

            if c.get("name"):
                clinics.append(c)

        return clinics

    return []


def get_all_city_slugs() -> list[str]:
    """Get city slugs from main /zubari/ page."""
    try:
        resp = requests.get(f"{BASE_URL}/__data.json", headers=HEADERS, timeout=15)
        data = resp.json()["nodes"][1]["data"]
    except Exception:
        return []

    slugs = []
    # Scan for slug-like strings followed by capitalized city names
    for i in range(len(data) - 1):
        val = data[i]
        nxt = data[i + 1] if i + 1 < len(data) else None
        if (isinstance(val, str) and isinstance(nxt, str)
                and val.islower() and not val.startswith("http")
                and nxt[0:1].isupper() and len(val) > 2 and " " not in val
                and val not in ("zubari", "true", "false", "high", "medium", "low", "green", "hzzo")):
            slugs.append(val)

    return list(dict.fromkeys(slugs))


def scrape_mojausluga() -> list[dict]:
    """Scrape all dental clinics."""
    print("\n=== Scraping mojausluga.hr ===")

    slugs = get_all_city_slugs()
    print(f"  Found {len(slugs)} city slugs from API")

    if len(slugs) < 50:
        print("  Adding fallback cities...")
        fallback = [
            "zagreb", "split", "rijeka", "osijek", "zadar", "slavonski-brod",
            "pula", "karlovac", "sisak", "varazdin", "dubrovnik", "bjelovar",
            "koprivnica", "vukovar", "velika-gorica", "samobor", "virovitica",
            "pozega", "cakovec", "krapina", "gospic", "zapresic", "solin",
            "kastela", "vinkovci", "kutina", "sibenik", "makarska", "sinj",
            "porec", "rovinj", "opatija", "crikvenica", "trogir", "sesvete",
            "dugo-selo", "ivanic-grad", "petrinja", "nova-gradiska", "dakovo",
            "nasice", "slatina", "ludbreg", "ivanec", "krizevci", "durdevac",
            "daruvar", "pakrac", "metkovic", "imotski", "knin", "drnis",
            "benkovac", "pag", "mali-losinj", "senj", "otocac", "delnice",
            "buzet", "pazin", "labin", "umag", "novigrad", "vodice",
        ]
        for f in fallback:
            if f not in slugs:
                slugs.append(f)

    all_clinics = []
    for i, slug in enumerate(slugs):
        print(f"  [{i+1}/{len(slugs)}] {slug}...", end=" ", flush=True)
        clinics = parse_city(slug)
        print(f"{len(clinics)}")
        all_clinics.extend(clinics)
        time.sleep(0.2)

    print(f"\n  Total scraped: {len(all_clinics)}")
    return all_clinics


def load_hzjz_dental() -> list[dict]:
    """Load dental entries from HZJZ Excel."""
    print(f"\n=== Loading HZJZ Excel ===")
    wb = openpyxl.load_workbook(HZJZ_EXCEL, read_only=True)
    clinics = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        cnt = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[0] or "").strip()
            if name and is_dental(name):
                clinics.append({
                    "name": name,
                    "county": str(row[1] or "").strip(),
                    "city": str(row[2] or "").strip(),
                    "address": str(row[3] or "").strip(),
                    "zip": str(row[4] or "").strip(),
                    "type": str(row[5] or "").strip(),
                })
                cnt += 1
        print(f"  {sheet}: {cnt}")
    wb.close()
    print(f"  Total: {len(clinics)}")
    return clinics


def normalize(name: str) -> set[str]:
    """Normalize a clinic name to a set of key words for matching."""
    n = name.lower()
    for rm in ["stomatološka ordinacija", "ordinacija dentalne medicine",
               "dental centar", "dentalni centar", "dentalna ordinacija",
               "dr.", "med.", "dent.", "mr.sc.", "spec.", "univ.", "mag.",
               "d.o.o.", ",", "."]:
        n = n.replace(rm, " ")
    return {w for w in n.split() if len(w) > 2}


def merge_and_export(hzjz: list[dict], scraped: list[dict]):
    """Merge datasets and export CSV."""
    print(f"\n=== Merging ===")

    # Build lookup: normalized name words -> scraped entry
    scraped_lookup = []
    for s in scraped:
        words = normalize(s.get("name", ""))
        if words:
            scraped_lookup.append((words, s))

    rows = []
    matched = 0

    for h in hzjz:
        row = {
            "naziv": h["name"],
            "zupanija": h["county"],
            "grad": h["city"],
            "adresa": h["address"],
            "postanski_broj": h["zip"],
            "vrsta": h["type"],
            "telefon": "",
            "email": "",
            "web": "",
            "hzzo_ugovor": "",
            "osiguranici": "",
            "vlasnistvo": "",
            "doktor": "",
            "izvor": "HZJZ",
        }

        h_words = normalize(h["name"])
        if not h_words:
            rows.append(row)
            continue

        best_match = None
        best_score = 0
        for s_words, s_data in scraped_lookup:
            overlap = len(h_words & s_words)
            union = len(h_words | s_words)
            score = overlap / union if union else 0
            if overlap >= 2 and score > best_score:
                best_score = score
                best_match = s_data

        if best_match and best_score >= 0.3:
            row["telefon"] = str(best_match.get("phone", "") or "")
            row["email"] = str(best_match.get("email", "") or "")
            row["web"] = str(best_match.get("websiteUrl", "") or "")
            row["hzzo_ugovor"] = "Da" if best_match.get("hzzoContract") else "Ne"
            row["osiguranici"] = str(best_match.get("insuredPatients", "") or "")
            row["vlasnistvo"] = str(best_match.get("ownership", "") or "")
            row["doktor"] = str(best_match.get("doctor_name", "") or "")
            row["izvor"] = "HZJZ + mojausluga.hr"
            matched += 1

        rows.append(row)

    # Add scraped-only entries
    hzjz_names = {normalize_simple(h["name"]) for h in hzjz}
    extra = 0
    for s in scraped:
        sn = normalize_simple(s.get("name", ""))
        if sn and sn not in hzjz_names:
            rows.append({
                "naziv": s.get("name", ""),
                "zupanija": "",
                "grad": s.get("cityName", s.get("city_slug", "").replace("-", " ").title()),
                "adresa": s.get("address", ""),
                "postanski_broj": "",
                "vrsta": s.get("ownership", ""),
                "telefon": str(s.get("phone", "") or ""),
                "email": str(s.get("email", "") or ""),
                "web": str(s.get("websiteUrl", "") or ""),
                "hzzo_ugovor": "Da" if s.get("hzzoContract") else "Ne",
                "osiguranici": str(s.get("insuredPatients", "") or ""),
                "vlasnistvo": s.get("ownership", ""),
                "doktor": str(s.get("doctor_name", "") or ""),
                "izvor": "mojausluga.hr",
            })
            extra += 1

    rows.sort(key=lambda r: (r["zupanija"], r["grad"], r["naziv"]))

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "naziv", "zupanija", "grad", "adresa", "postanski_broj", "vrsta",
            "telefon", "email", "web", "hzzo_ugovor", "osiguranici",
            "vlasnistvo", "doktor", "izvor",
        ])
        writer.writeheader()
        writer.writerows(rows)

    with_phone = sum(1 for r in rows if r["telefon"])
    print(f"  HZJZ dental entries: {len(hzjz)}")
    print(f"  Matched with contacts: {matched}")
    print(f"  Extra from scrape: {extra}")
    print(f"  Total rows: {len(rows)}")
    print(f"  With phone: {with_phone} ({100*with_phone//max(len(rows),1)}%)")
    print(f"\n  Output: {OUTPUT_CSV}")


def normalize_simple(name: str) -> str:
    return " ".join(name.lower().split())


def main():
    print("=" * 60)
    print("  HM Digital — Dental Clinic Contact Scraper")
    print("=" * 60)
    hzjz = load_hzjz_dental()
    scraped = scrape_mojausluga()
    merge_and_export(hzjz, scraped)
    print("\nDone!")


if __name__ == "__main__":
    main()
